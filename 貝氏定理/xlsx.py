import os
from openpyxl import Workbook
from openpyxl import load_workbook

class Xlsx():
        
    def __init__(self, name = 'new'):
     
        self.name = name
        
        if os.path.exists(name):
            self.wb = load_workbook(name)
        else:
            self.wb = Workbook(name)
            
            self.ws = self.wb.create_sheet()
            self.ws.title = '臺北市'
            self.ws.append(['loc', 'time', 'temp'])
       
            
        
        

    def loadSheet(self, sheet = 'sheet'):
        
        sh = self.wb[sheet]
            
        return sh


    def Sheetwrite(self,sheet,row):
# 指定某个sheet中添加内容
        a = self.loadSheet(sheet)
        a.append(row)
    
    def save(self):
    # 保存数据到文件中
        self.wb.save(self.name)
    
    
    def read(self,sheet):
        
        now = datetime.datetime.now()
        hour = datetime.timedelta(0,0,0,0,10)
        sh = self.loadSheet(sheet)
        
        rows = sh.rows
        #content = []
    

        for row in rows:
            try:
                if abs(now-row[3].value)<= hour:
                        print('已輸入')
                        return True
            except TypeError:
                pass
        
        
        #print(write, content)
        return False
    
