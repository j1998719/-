#import time
import datetime
import os
import openpyxl
#import re
from openpyxl import Workbook
from openpyxl import load_workbook
import math
from history_data import *

url='https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
hourly_data= '預測資料.xlsx'
observ = '觀測資料.xlsx'
URL={
    '臺北市':'https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
    
    }
fifty_mins = datetime.timedelta(0,0,0,0,16)
five_mins = datetime.timedelta(0,0,0,0,5)

test_month = 1


class Xlsx():
        
    def __init__(self, file=hourly_data):
         
        self.file = file

        if os.path.exists(file):
            self.wb = load_workbook(file)
        else:
            self.wb = Workbook(file)

            self.ws = self.wb.create_sheet()
            self.ws.title = '臺北市'
            self.ws.append(['time', 'temp'])
                   

    def loadSheet(self, sheet = 'sheet'):
        
        sh = self.wb[sheet]
            
        return sh


    def Sheetwrite(self,sheet,row):
        a = self.loadSheet(sheet)
        a.append(row)
    
    def save(self):
        self.wb.save(self.file)
    
    
    def read(self,sheet, write):
        
        timedelta = datetime.timedelta(0,0,0,0,1)
        sh = self.loadSheet(sheet)
        
        rows = sh.rows
        content = []
    

        for row in rows:
            content.append([col.value for col in row])
        
        for row in content:
            if not type(row[0]) == str: 
                if abs(row[0] - write[0]) <= timedelta and row[1] == write[1]:
                    return True
            
        
        return False
    

 ############################################################################################################################################       


hourly = Xlsx(hourly_data).loadSheet('臺北市')
Observ = Xlsx(observ).loadSheet('臺北市')




def is_number(x):
    try:
        a = float(x)
        return True
    
    except:
        pass
        



#condition probability
conp = [[0,0,0],
     [0,0,0],
     [0,0,0]
     ]
#condition
con = [0,0,0]


def scope(load, i=1):
    #資料溫度得範圍
    Max=0
    Min=100
    
    for row in load.rows:
        if is_number(row[i].value):
            if float(row[i].value)!= -99:
                if row[0].value != '大屯山' and row[0].value != '陽明山' and row[0].value != '鞍部':
                    if float(row[i].value) > Max:
                        Max=float(row[i].value)
                    if float(row[i].value) < Min:
                        Min=float(row[i].value)
                    
    return [Min, Max]




def low_high():
    #兩份資料的最低最高
    H_scope = scope(hourly)
    O_scope = scope(Observ,2)

    if H_scope[0] < O_scope[0]:
        a = H_scope[0]
    else:
        a = O_scope[0]

    if H_scope[1] > O_scope[1]:
        b = H_scope[1]
    else:
        b = O_scope[1]

    return [math.floor(a),math.ceil(b)]


#high跑出76，好大
low = low_high()[0]
high = low_high()[1]+1

if high >50:
    high = 50


#找到資料的區間，三等分分為偏低 中間 偏高
def gap():

    delta = high - low

    gap1 = low + float(delta)*0.3
    gap2 = low + float(delta)*0.7

    return [gap1, gap2]


def GAP(x=gap()[0], y=gap()[1]):

    b = [0,x,y,100]
    print(b)
    
    for hr in hourly.rows:
        for obs in Observ.rows:
            
            if hr[0].value != 'time' and obs[1].value!='time' and hr[0].value != 'X' and obs[1].value!='X':
                    

                if abs(hr[0].value-obs[1].value) < fifty_mins:

                    for i in range(0,3):
                        if b[i] < float(hr[1].value) < b[i+1]:
                            con[i] += 1
                            
                            for j in range(0,3):
                                if b[j] < float(obs[2].value) < b[j+1]:
                                    conp[i][j] += 1
                                
                                
                        



            
    t = ['偏低', '中間', '偏高']                

    for i in range(0,3):
        for j in range(0,3):

            if con[i]!= 0:
                text = '預測:%s 觀察:%s ==> %f'%(t[i], t[j], float(conp[i][j]/con[i]))
            else:
                text = '預測:%s 觀察:%s ==> %f'%(t[i], t[j], float(0))

            print(text)


def Bayes(x=15):
    #預測溫度是X的情況，觀測溫度是其他溫度的機率

    #pre_prob = comfirm(0,x)[3]
    prob = 0
    temp = -99
    
    for i in range(low,high):
        

        t = SimpleWay(i, x)[2]


        if t > prob:
            prob = t
            temp=i

    
    
    if temp!=-99:
        simpleway = SimpleWay(temp, x)[2]
        origin_prob = SimpleWay(x,x)[2]
        print('temp:',x,'origin_prob',origin_prob,'expect:',temp,'prob:', prob,'#:',simpleway)
    else:
        origin_prob = SimpleWay(x,x)[2]
        print('temp:',x,'origin_prob',origin_prob,'expect:','NoneData','prob:', prob)


def comfirm():
    #A=實際溫度 B=預測溫度 P(B|A)
    
    dic = {}
    DIC = {}
    
    for i in range(low,high):
        DIC[str(i)] = 0
        
    for i in range(low,high):
        dic[str(i)] = {}
        
        for j in range(low,high):
            dic[str(i)][str(j)] = 0
    
    for hr in hourly.rows:
        for obs in Observ.rows:

            if hr[0].value != 'time' and obs[1].value!='time' and hr[0].value != 'X' and obs[1].value!='X' and obs[1].value!='-99':
                if int(hr[0].value.month) == test_month :
                    #如果讓資料僅限在前後一個月
                    if abs(hr[0].value-obs[1].value) < fifty_mins:

                        check = 0
                        for i in range(low,high):
                            if check != 1:
                                 
                                for j in range(low,high):

                                    if  abs(float(obs[2].value)-i)<0.5:
                                        DIC[str(i)] += 1
                                        
                                        if  abs(float(hr[1].value)-j)<0.5:
                                            dic[str(i)][str(j)] += 1
                                            check = 1
                                            break
                            else:
                                break

            
                

    for i in range(low,high):
        for j in range(low,high):

            if DIC[str(i)] !=0:
                dic[str(i)][str(j)] /= DIC[str(i)]

    #dic[A][B] = P(B|A)
    return dic



def adjust():

    # P(A|B) = P(A)P(B|A)/P(An)P(B|An)
    
    his_dic = month(test_month)
    pre_dic = comfirm()
    
    
    dic = {}
    DIC = {}
    
    for i in range(low,high):
        DIC[str(i)] = 0
        
    for i in range(low,high):
        dic[str(i)] = {}
        
        for j in range(low,high):
            dic[str(i)][str(j)] = 0

    #DIC[B] = P(B)    
    for i in range(low, high):
        for j in range(low, high):
            DIC[str(i)] +=  his_dic[str(j)] * pre_dic[str(j)][str(i)]
    
    #dic[B][A] = P(A|B)
    for i in range(low,high):
        for j in range(low, high):
            if DIC[str(i)] != 0:
                dic[str(i)][str(j)] = his_dic[str(j)]*pre_dic[str(j)][str(i)] / DIC[str(i)]
                if dic[str(i)][str(j)] > 1 :
                    print(f'error: dic[{i}][{J}] = {dic[str(i)][str(j)]}')
            

    return dic


def SimpleWay():
    #(直接統計的方法)在預測溫度是Y的情況，觀測資料是X的機率
    #P(A|B)

    dic = {}
    DIC = {}
    
    for i in range(low,high):
        DIC[str(i)] = 0
        
    for i in range(low,high):
        dic[str(i)] = {}
        
        for j in range(low,high):
            dic[str(i)][str(j)] = 0
    
    for hr in hourly.rows:
        for obs in Observ.rows:

            if hr[0].value != 'time' and obs[1].value!='time' and hr[0].value != 'X' and obs[1].value!='X' and obs[1].value != '-99':
                if int(hr[0].value.month) == month:
                    if abs(hr[0].value-obs[1].value) < fifty_mins:

                        check = 0
                        for i in range(low,high):
                            if check != 1:
                                 
                                for j in range(low,high):

                                    if  abs(float(hr[1].value)-i)<0.5:
                                        DIC[str(i)] += 1
                                        
                                        if  abs(float(obs[2].value)-j)<0.5:
                                            dic[str(i)][str(j)] += 1
                                            check = 1
                                            break
                            else:
                                break

            
                

    for i in range(low,high):
        for j in range(low,high):

            if DIC[str(i)] !=0:
                dic[str(i)][str(j)] /= DIC[str(i)]

    #dic[B][A] = P(A|B)
    return dic


adj = adjust()

def adj_temp():

    #給定預測溫度X，找出實際最有可能發生的溫度
    
    biggest = 0
    biggest_no = 0

    dic = {}
    for key in adj:

        biggest = 0
        biggest_no = 0
    
        for i in range(low, high):
            if adj.get(key).get(str(i)) > biggest:
                
                biggest = adj[key][str(i)]
                biggest_no = i

        dic[key] = biggest_no

    
    
    return dic

adj_temp = adj_temp()

def main():

    count = 0

    pre_right = [0,0,0,0]
    post_right = [0,0,0,0]
    
    delta = 0
    
    for hr in hourly.rows:
        if hr[0].value != 'time' and hr[0].value != 'X':
            
            if int(hr[0].value.month) == test_month:
                for obs in Observ.rows:
                    
                    if obs[1].value!='time' and obs[1].value!='X' and obs[1].value != '-99':
                        if abs(obs[1].value - hr[0].value) < fifty_mins:

                            adjust = adj_temp[str(hr[1].value)]
                            
##                            if abs(float(hr[1].value) - float(obs[2].value)) >=4:
##                                print(f'預測時間:{hr[0].value}\t觀測時間:{obs[1].value}')
##                                print(f'預測溫度:{float(hr[1].value)} 調整後溫度:{adjust} 實際觀測溫度:{obs[2].value}')
##                                print('==============================================================================')
##                                
                            count += 1
                            delta += math.pow(abs(adjust - float(obs[2].value)),2)

                            for i in range(0,4):
                                if abs(adjust - float(obs[2].value)) <= (i+1):
                                    post_right[i] += 1
                                if abs(float(hr[1].value) - float(obs[2].value)) <= (i+1):
                                    pre_right[i] += 1
                                

    if count != 0:
                                      
        for i in range(0,4):
            pre_right[i] /= count
            post_right[i] /= count
                                      
        delta /= math.sqrt(delta/count)
        
    print('先Pre後Post')
    for i in range(0,4):
        print('正負',i+1,":",pre_right[i],post_right[i])
            

##for i in range(low, high):
##    print(f'預測:{i}\t最可能:{adj_temp(i)}')


main()

Xlsx(hourly_data).save()
Xlsx(observ).save()






