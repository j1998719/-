
# coding: utf-8

# In[1]:


from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import urllib
import datetime
import os
import openpyxl
import re
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

url='https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
hourly_data= '預測資料.xlsx'
observ = '觀測資料.xlsx'
URL={
    '臺北市':'https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
    
    }
year=datetime.datetime.now().year
month=datetime.datetime.now().month
day=datetime.datetime.now().day
now=datetime.datetime.now()
oneday = datetime.timedelta(1)
# In[2]:


class Xlsx():
        
    def __init__(self, file=hourly_data):
         
        self.file = file

        if os.path.exists(file):
            self.wb = load_workbook(file)
        else:
            self.wb = Workbook(file)

            self.ws = self.wb.create_sheet()
            self.ws.title = '臺北市'
            self.ws.append(['time', 'temp'])
                   

    def loadSheet(self, sheet = 'sheet'):
        
        sh = self.wb[sheet]
            
        return sh


    def Sheetwrite(self,sheet,row):
# 指定某个sheet中添加内容
        a = self.loadSheet(sheet)
        a.append(row)
    
    def save(self):
    # 保存数据到文件中
        self.wb.save(self.file)
    
    
    def read(self,sheet, write):
        
##        one_min = datetime.timedelta(0,0,0,0,1)
##        ten_min = datetime.timedelta(0,0,0,0,10)
##
##        sh = self.loadSheet(sheet)
##        
##        rows = sh.rows
##        content = []
##    
##
##        for row in rows:
##            content.append([col.value for col in row])
##        
##        for row in content:
##            #print(row)
##            if not type(row[0]) == str: 
##                if abs(row[0] - write[0]) <= ten_min and abs(row[2] - write[2])<ten_min:
##                    return True
##            
        
        #print(write, content)
        return False
    

        

hourly = Xlsx(hourly_data)
hourly.save()


# In[6]:


def soup(url):
    driver= webdriver.Chrome()
    driver.get(url)
    soup = BeautifulSoup(driver.page_source,'lxml')
    driver.quit()
    
    return soup

def source(url):


    if day < 10:
        DAY = '0'+str(day)
    else:
        DAY = str(day)
        
    Soup = soup(url)
    
    trs = Soup.find_all( 'tr')

    i=0
    for tr in trs:
        span = tr.find('span',{'class':"dsx-date"})
        td = tr.find('td', {'class':'temp'})
        try:
            
            text=str(year)+str(month)+DAY+span.text
            strpt= datetime.datetime.strptime(text, '%Y%m%d%H:%M')
            
            if strpt < now:
##                try:
##                    if day+1 <10:
##                        DAY = '0'+str(day+1)
##                    else:
##                        DAY = str(day+1)
##                        
##                    text=str(year)+str(month)+DAY+span.text
##                    strpt= datetime.datetime.strptime(text, '%Y%m%d%H:%M')
##                    
##                except ValueError:
##                    
##                    if month<12:
##                        
##                        text=str(year)+str(month+1)+'01'+span.text
##                        strpt= datetime.datetime.strptime(text, '%Y%m%d%H:%M')
##                    else:
##                        test=str(year+1)+'01'+'01'+span.text
                strpt += oneday


            temp = td.span.text[0:2]
   
            hourly.Sheetwrite('臺北市',[strpt,temp])
            print('輸入: ', [strpt,temp])
            i=i+1
                
                
            
                
                
        except AttributeError:
            print('pass')

    print('輸入',i,'筆資料')
    print('end')



##while 1:
##        now = time.ctime(int(time.time()))
##        print("the now time is " + str(now) )
##        hourly = Xlsx(hourly_data)
##        source(url)
##        hourly.save()
##        for i in range(0,59):
##            now = time.ctime(time.time())
##            print("the now time is " + str(now) )
##            time.sleep(60)

if __name__ == '__main__':
    hourly = Xlsx(hourly_data)
    source(url)
    hourly.save()
    print("預測執行完畢")
    time.sleep(10)
    


# In[104]:




