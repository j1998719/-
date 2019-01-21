
# coding: utf-8

# In[1]:


from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import urllib
import datetime
import os
import openpyxl
import re
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


# In[2]:


city_url = {'基隆市':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46692&idx=10017#ui-tabs-1',
           '臺中市':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46749&idx=66#ui-tabs-2',
           '臺南市':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46741&idx=67#ui-tabs-3',
           '宜蘭縣':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46699&idx=10002#ui-tabs-4',
           '花蓮縣':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46699&idx=10015#ui-tabs-4',
           '臺東縣':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46699&idx=10014#ui-tabs-4',
           '澎湖縣':'https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46799&idx=10016#ui-tabs-5'
           }

CITY = {'基隆市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Keelung_City.htm',
               '臺北市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Taipei_City.htm',
               '新北市':'https://www.cwb.gov.tw/V7/forecast/taiwan/New_Taipei_City.htm',
               '桃園市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Taoyuan_City.htm',
               '新竹市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Hsinchu_City.htm',
               '新竹縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Hsinchu_County.htm',
               '苗栗縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Miaoli_County.htm',
           '臺中市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Taichung_City.htm',
           '彰化縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Changhua_County.htm',
           '南投縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Nantou_County.htm',
           '雲林縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Yunlin_County.htm',
           '嘉義市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Chiayi_City.htm',
           '嘉義縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Chiayi_County.htm',
           '臺南市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Tainan_City.htm',
           '高雄市':'https://www.cwb.gov.tw/V7/forecast/taiwan/Kaohsiung_City.htm',
           '屏東縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Pingtung_County.htm',
           '宜蘭縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Yilan_County.htm',
           '花蓮縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Hualien_County.htm',
           '臺東縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Taitung_County.htm',
           '連江縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Lienchiang_County.htm',
           '金門縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Kinmen_County.htm',
           '澎湖縣':'https://www.cwb.gov.tw/V7/forecast/taiwan/Penghu_County.htm'
           }
observ = '觀測資料.xlsx'
hourly_data = '預測資料.xlsx'
year = datetime.datetime.now().year
now = datetime.datetime.now()


# In[3]:


class Xlsx():
        
    def __init__(self, observ):
     

        if os.path.exists(observ):
            self.wb = load_workbook(observ)
        else:
            self.wb = Workbook(observ)

            for key in CITY:
                self.ws = self.wb.create_sheet()
                self.ws.title = key
                self.ws.append(['loc', 'time', 'temp', 'run_time'])
##            self.ws = self.wb.create_sheet()
##            self.ws.title = '臺北市'
##            self.ws.append(['loc', 'time', 'temp', 'run_time'])
##            
            
        
        

    def loadSheet(self, sheet = 'sheet'):
        
        sh = self.wb[sheet]
            
        return sh


    def Sheetwrite(self,sheet,row):
# 指定某个sheet中添加内容
        a = self.loadSheet(sheet)
        a.append(row)
    
    def save(self):
    # 保存数据到文件中
        self.wb.save(observ)
    
    
    def read(self,sheet):
        
        now = datetime.datetime.now()
        hour = datetime.timedelta(0,0,0,0,10)
        sh = self.loadSheet(sheet)
        
        rows = sh.rows
        #content = []
    

        for row in rows:
            try:
                if abs(now-row[3].value)<= hour:
                        print('已輸入')
                        return True
            except TypeError:
                pass
        
        
        #print(write, content)
        return False
    


# In[4]:


#html抓取 觀測站 觀測時間 溫度 寫入時間

def TR(tr):
    td = tr.find_all('td')
    strpt = str(year)+td[2].text
    try:
        
        if td[3].text =='-':
            temp='-99'
        else:
            temp = td[3].text 

    except IndexError:
        temp = 'X'
    
    loc = td[1].text

    try:
        strptime = datetime.datetime.strptime(strpt, '%Y%m/%d %H:%M')
    except ValueError:
        strptime = "X"     
        
    return [loc, strptime, temp, now]
                


# In[7]:


observation = Xlsx(observ)
observation.save()

def source():
    driver = webdriver.Chrome()
    #selenium.webdriver抓取動態網頁
    

    observation = Xlsx(observ)

    
    content = []
    driver.get('https://www.cwb.gov.tw/V7/observe/real/NewObs.htm?Station=46692&idx=10017#ui-tabs-1')
    soup = BeautifulSoup(driver.page_source, features='lxml')
    id63s = soup.find_all( 'table', {"class": "tablesorter"})
    #tablesorter取得觀測圖表
    citynames = driver.find_elements_by_xpath("//a[@href='#self']")
    city_num = len(citynames)
    city = []
    for c in citynames:
        city.append(c.text)
        #取得城市名稱
    
    #print(id63)
    i = 0
    #判斷資料是否重複抓取
    try:
        if not observation.read(city[i]):
            for id63 in id63s:
                #print(city[i])
                for tr in id63.tbody.find_all('tr'):
                    #html抓取 觀測站 觀測時間 溫度 寫入時間
                    td = TR(tr)
                    observation.Sheetwrite(city[i], td)    
                    #寫入資料

                i = i+1
                #同一個網頁有數個城市資料，用迴圈
    except IndexError:
        pass
        
    driver.quit()
    observation.save()
    print('END')

# In[8]:

##while 1:
##    
##    source()
##    for i in range(0,59):
##            now = time.ctime(time.time())
##            print("the now time is " + str(now) )
##            time.sleep(60)

if __name__ == '__main__':
    source()
    print("觀測執行完畢")
    time.sleep(10)
