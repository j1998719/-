import time
import datetime
import os
import openpyxl
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import math

url='https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
hourly_data= '預測資料.xlsx'
observ = '觀測資料.xlsx'
URL={
    '臺北市':'https://weather.com/zh-TW/weather/hourbyhour/l/Taipei+Taiwan+TWXX0021:1:TW'
    
    }
fifty_mins = datetime.timedelta(0,0,0,0,16)
five_mins = datetime.timedelta(0,0,0,0,5)


class Xlsx():
        
    def __init__(self, file=hourly_data):
         
        self.file = file

        if os.path.exists(file):
            self.wb = load_workbook(file)
        else:
            self.wb = Workbook(file)

            self.ws = self.wb.create_sheet()
            self.ws.title = '臺北市'
            self.ws.append(['time', 'temp'])
                   

    def loadSheet(self, sheet = 'sheet'):
        
        sh = self.wb[sheet]
            
        return sh


    def Sheetwrite(self,sheet,row):
        a = self.loadSheet(sheet)
        a.append(row)
    
    def save(self):
        self.wb.save(self.file)
    
    
    def read(self,sheet, write):
        
        timedelta = datetime.timedelta(0,0,0,0,1)
        sh = self.loadSheet(sheet)
        
        rows = sh.rows
        content = []
    

        for row in rows:
            content.append([col.value for col in row])
        
        for row in content:
            if not type(row[0]) == str: 
                if abs(row[0] - write[0]) <= timedelta and row[1] == write[1]:
                    return True
            
        
        return False
    

        


hourly = Xlsx(hourly_data).loadSheet('臺北市')
Observ = Xlsx(observ).loadSheet('臺北市')




def is_number(x):
    try:
        a = float(x)
        return True
    
    except ValueError:
        return False

    except TypeError:
        return False
        

def comfirm(x=15,y=15):
    #在預測溫度是Y的情況，觀測資料是X的機率
    i=0.0
    total=0.0
    pre = 0.0
    pre_total=0
    
    for hr in hourly.rows:
        for obs in Observ.rows:
            try:

                    
                if abs(hr[0].value-obs[1].value) < fifty_mins:
                    total+=1
                    
                    if  abs(int(hr[1].value)-y)<0.5:
                        pre+=1
                        
                        if  abs(float(obs[2].value)-x)<0.5:
                            i+=1
            except:
                pass
            
                

    if total!=0:
        prob = round(i/total,3)
    else:
        prob = 0

    if total!=0:
        pre_prob = round(pre/total,3)
    else:
        pre_prob = 0

    return [total,i,prob,pre_prob]



def SimpleWay(x=15,y=15):
    #(直接統計的方法)在預測溫度是Y的情況，觀測資料是X的機率
    i=0.0
    total=0.0
    for hr in hourly.rows:
        try:
            if abs(float(hr[1].value)-y)<0.5:
                for obs in Observ.rows:
                    
                    try:
                        if abs(hr[0].value - obs[1].value) < fifty_mins:
                            total+=1
                            
                            if abs(float(obs[2].value)-x)<0.5:
                                i+=1
                    except:
                        pass
                    
        except:
            pass
            
                

    if total!=0:
        prob = round(i/total,3)
    else:
        prob = 0

    return [total,i,prob]



def scope(load, i=1):
    #資料溫度得範圍
    Max=0
    Min=100
    for row in load.rows:
        if is_number(row[i].value):
            if float(row[i].value)!= -99:
                if float(row[i].value) > Max:
                    Max=float(row[i].value)
                if float(row[i].value) < Min:
                    Min=float(row[i].value)
    return [Min, Max]


def low_high():
    #兩份資料的最低最高
    H_scope = scope(hourly)
    O_scope = scope(Observ,2)

    if H_scope[0] < O_scope[0]:
        a = H_scope[0]
    else:
        a = O_scope[0]

    if H_scope[1] > O_scope[1]:
        b = H_scope[1]
    else:
        b = O_scope[1]

    return [math.floor(a),math.ceil(b)]

def Bayes(x=15):
    #預測溫度是X的情況，觀測溫度是其他溫度的機率
    low=low_high()[0]
    high=low_high()[1]+1
    #pre_prob = comfirm(0,x)[3]
    prob = 0
    temp = -99
    
    for i in range(low,high):
        
        t = SimpleWay(i, x)[2]


        if t > prob:
            prob = t
            temp=i

    
    
    if temp!=-99:
        simpleway = SimpleWay(temp, x)[2]
        origin_prob = SimpleWay(x,x)[2]
        print('temp:',x,'origin_prob',origin_prob,'expect:',temp,'prob:', prob,'#:',simpleway)
    else:
        origin_prob = SimpleWay(x,x)[2]
        print('temp:',x,'origin_prob',origin_prob,'expect:','NoneData','prob:', prob)


    
def gap():
    low = low_high()[0]
    high = low_high()[1]

    delta = high - low

    gap1 = low + float(delta)*0.3
    gap2 = low + float(delta)*0.7

    return [gap1, gap2]

def GAP(x=gap()[0], y=gap()[1]):

    b = [0,x,y,100]
    print(b)
    
    for hr in hourly.rows:
        for obs in Observ.rows:
            
            if hr[0].value != 'time' and obs[1].value!='time' and hr[0].value != 'X' and obs[1].value!='X':
                    

                if abs(hr[0].value-obs[1].value) < fifty_mins:

                    for i in range(0,3):
                        if b[i] < float(hr[1].value) < b[i+1]:
                            con[i] += 1
                        for j in range(0,3):
                            if b[i] < float(hr[1].value) < b[i+1] and b[j] < float(obs[2].value) < b[j+1]:
                                conp[i][j] += 1
                                
                                
                        



            
    t = ['偏低', '中間', '偏高']                

    for i in range(0,3):
        for j in range(0,3):

            if con[i]!= 0:
                text = '預測:%s 觀察:%s ==> %f'%(t[i], t[j], float(conp[i][j]/con[i]))
            else:
                text = '預測:%s 觀察:%s ==> %f'%(t[i], t[j], float(0))

            print(text)










